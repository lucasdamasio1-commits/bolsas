"""
Monitor de Bolsas de Pesquisa — v5 (Precisão de Links Individuais)
====================================
MUDANÇA PRINCIPAL: Inclusão de um extrator cirúrgico para portais 
agregadores como o Bertha (CAPES). O script isola cada oportunidade 
em sua respectiva linha no Excel, extraindo o link EXATO de cada edital.

Áreas monitoradas: Administração, Marketing, IA/Dados,
                   Empreendedorismo, Inovação, Economia/Finanças.
Níveis: Mestrado, Doutorado, Pós-doc, Pesquisador Sênior.

Fontes:
  Brasil   → CAPES (lista principal + portal Bertha), CNPq, FAPESP
  Europa   → UKRI, MSCA, CORDIS, OpenAire, DAAD, Erasmus+
  N.América→ Fulbright, NIH Reporter
  Portugal → FCT
  Global   → ScholarshipPortal, FindAPhD

Dependências:
    pip install requests beautifulsoup4 openpyxl feedparser lxml
"""

import hashlib
import html as html_lib
import json
import logging
import os
import shutil
import subprocess
import time
from datetime import date
from urllib.parse import urljoin, urlparse

import feedparser
import openpyxl
<<<<<<< HEAD
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import os
import logging
import subprocess
import re
from datetime import date
from urllib.parse import urljoin

=======
>>>>>>> f9770e4 (Atualiza pagina de bolsas no GitHub Pages)
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

PASTA_SAIDA   = r"C:\Users\lucas\OneDrive\Doutorado\Scripts\bolsas"
ARQUIVO_EXCEL = os.path.join(PASTA_SAIDA, "bolsas_pesquisa.xlsx")
ARQUIVO_HTML  = os.path.join(PASTA_SAIDA, "index.html")
PASTA_GITHUB_PAGES = os.path.join(PASTA_SAIDA, "bolsas")
ARQUIVO_EXCEL_PAGES = os.path.join(PASTA_GITHUB_PAGES, "bolsas_pesquisa.xlsx")
ARQUIVO_HTML_PAGES = os.path.join(PASTA_GITHUB_PAGES, "index.html")
ARQUIVO_IDS_PAGES = os.path.join(PASTA_GITHUB_PAGES, "ids_vistos.json")
ARQUIVO_LOG   = os.path.join(PASTA_SAIDA, "coleta.log")
ARQUIVO_IDS   = os.path.join(PASTA_SAIDA, "ids_vistos.json")

TIMEOUT             = 20
PAUSA_ENTRE_FONTES  = 2   

# ─────────────────────────────────────────────────────────────────────────────
# PALAVRAS-CHAVE DAS SUAS ÁREAS
# ─────────────────────────────────────────────────────────────────────────────

AREAS_INTERESSE = {
    # administração / gestão
    "administracao", "gestao", "management", "organizational",
    "governance", "lideranca", "leadership", "strategy",
    # marketing / comportamento
    "marketing", "consumer", "comportamento", "consumidor",
    "brand", "advertising", "digital marketing", "comunicacao",
    # IA / dados
    "artificial intelligence", "machine learning", "deep learning",
    "data science", "big data", "algoritmo", "nlp",
    "inteligencia artificial", "dados", "analytics",
    # empreendedorismo / inovação
    "entrepreneurship", "empreendedorismo", "inovacao", "innovation",
    "startup", "venture", "ecossistema", "spin-off",
    # economia / finanças
    "economia", "economics", "finance", "financas", "fintech",
    "mercado", "market", "investimento", "fiscal",
    # termos acadêmicos gerais
    "phd", "doctoral", "doutorado", "mestrado", "master",
    "postdoc", "fellowship", "scholarship", "grant", "bolsa",
    "pesquisa", "research", "funding", "edital", "call",
}

def area_relevante(texto: str) -> bool:
    import unicodedata
    t = unicodedata.normalize("NFD", texto.lower()).encode("ascii", "ignore").decode()
    return any(p in t for p in AREAS_INTERESSE)

def detectar_nivel(texto: str) -> str:
    t = texto.lower()
    if any(x in t for x in ("phd", "doctoral", "doutorado")):
        return "Doutorado"
    if any(x in t for x in ("master", "mestrado")):
        return "Mestrado"
    if any(x in t for x in ("postdoc", "pos-doutorado", "senior researcher", "pesquisador senior")):
        return "Pos-doc / Senior"
    return "Geral / Multiplos"

# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS HTTP E HEURÍSTICA DE LINKS
# ─────────────────────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
})

def get_html(url: str) -> BeautifulSoup | None:
    try:
        r = SESSION.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return BeautifulSoup(r.text, "lxml")
    except Exception as e:
        log.warning(f"HTTP erro [{url[:60]}]: {e}")
        return None

def get_json(url: str, **kwargs) -> dict | list | None:
    try:
        r = SESSION.get(url, timeout=TIMEOUT, **kwargs)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        log.warning(f"JSON erro [{url[:60]}]: {e}")
        return None

def abs_url(href: str, base: str) -> str | None:
    if not href:
        return None
    href = href.strip()
    if href.startswith(("javascript:", "#", "mailto:")):
        return None
    if href.startswith("http"):
        return href
    p = urlparse(base)
    if href.startswith("/"):
        return f"{p.scheme}://{p.netloc}{href}"
    return urljoin(base, href)

def encontrar_melhor_link(elemento, base_url: str) -> str:
    links = elemento.find_all("a", href=True)
    if not links:
        return base_url

    melhor_link = None
    pontuacao_maxima = -1

    for a in links:
        href = a.get("href", "").lower()
        texto = a.get_text(strip=True).lower()
        
        if href.startswith(("javascript", "mailto", "#")) or any(x in href for x in ["twitter.com", "facebook.com", "linkedin.com"]):
            continue

        pontuacao = 0
        if href.endswith((".pdf", ".doc", ".docx")) or "download" in href:
            pontuacao = 100
        elif any(p in href for p in ["edital", "chamada", "call", "guidelines", "apply", "candidatura"]):
            pontuacao = 80
        elif any(p in texto for p in ["edital", "chamada", "call", "apply now", "candidatar", "inscreva-se", "leia mais", "acessar"]):
            pontuacao = 70
        elif a.find_parent(["h2", "h3", "h4"]):
            pontuacao = 60
        else:
            pontuacao = 10

        if pontuacao > pontuacao_maxima:
            pontuacao_maxima = pontuacao
            melhor_link = abs_url(a["href"], base_url)

    return melhor_link or base_url

def bolsa(titulo, fonte, pais, nivel, prazo, link, area="Multiplas"):
    return {
        "titulo": titulo.strip(),
        "fonte":  fonte,
        "pais":   pais,
        "area":   area,
        "nivel":  nivel or detectar_nivel(titulo),
        "prazo":  prazo or "Consultar edital",
        "link":   link or "",
        "data":   str(date.today()),
    }

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────

os.makedirs(PASTA_SAIDA, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(ARQUIVO_LOG, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ═════════════════════════════════════════════════════════════════════════════
# COLETORES ESPECÍFICOS POR FONTE
# ═════════════════════════════════════════════════════════════════════════════

# ── BRASIL ───────────────────────────────────────────────────────────────────

def coletar_capes() -> list[dict]:
    resultados = []
    urls = [
        "https://www.gov.br/capes/pt-br/acesso-a-informacao/acoes-e-programas/bolsas/bolsas-no-pais",
        "https://www.gov.br/capes/pt-br/acesso-a-informacao/acoes-e-programas/bolsas/bolsas-no-exterior",
        "https://www.gov.br/capes/pt-br/acesso-a-informacao/acoes-e-programas/avaliacao/editais",
    ]
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for item in soup.select("article, .tileItem, .summary, li.edital"):
            h = item.find(["h2", "h3", "h4", "a"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 12 or not area_relevante(titulo):
                continue
            
            link = encontrar_melhor_link(item, url)
            
            prazo = ""
            desc = item.get_text(" ", strip=True)
            import re
            m = re.search(r"\d{2}/\d{2}/\d{4}", desc)
            if m:
                prazo = m.group()
            resultados.append(bolsa(titulo, "CAPES", "Brasil", detectar_nivel(titulo), prazo, link))

    log.info(f"CAPES Gov.br → {len(resultados)} editais")
    return resultados

def coletar_capes_bertha() -> list[dict]:
    """
    Coletor específico para o portal CAPES Bertha.
    Cada card/link de oportunidade vira uma linha própria, preservando o link
    individual da oportunidade, inclusive quando o destino final está no gov.br.
    """
    import re

    resultados = []
    url_base = "https://bertha.capes.gov.br/oportunidades"
    api_url = "https://bertha.capes.gov.br/api/editais-abertos"

    data = get_json(api_url, headers={"Accept": "application/json"})
    if isinstance(data, dict):
        oportunidades_processadas_api = set()
        for item in data.get("items", []):
            if not isinstance(item, dict):
                continue

            programa = (item.get("programa_nome") or item.get("programa_sigla") or "").strip()
            edital = (item.get("edital_nome") or item.get("edital_sigla") or "").strip()
            partes_titulo = [p for p in (programa, edital) if p]
            titulo = " - ".join(dict.fromkeys(partes_titulo))
            if not titulo:
                continue

            link = (item.get("url") or "").strip()
            if not link and item.get("edital_sigla"):
                link = f"https://inscricao.capes.gov.br/edital/{item['edital_sigla']}"
            if not link:
                link = url_base

            chave_oportunidade = "::".join([
                str(item.get("edital_sigla") or "").strip().lower(),
                str(item.get("programa_sigla") or "").strip().lower(),
                titulo.lower(),
                link.lower(),
            ])
            if chave_oportunidade in oportunidades_processadas_api:
                continue

            prazo = (item.get("termino_inscricao") or "").strip().split(" ")[0]
            modalidades = ", ".join(item.get("modalidades_bolsa") or [])
            texto_nivel = f"{titulo} {modalidades}"
            area = (item.get("tipo_inscricao") or "CAPES Bertha").title()

            resultados.append(bolsa(
                titulo,
                "CAPES (Bertha)",
                "Brasil",
                detectar_nivel(texto_nivel),
                prazo,
                link,
                area=area,
            ))
            oportunidades_processadas_api.add(chave_oportunidade)

        if resultados:
            log.info(f"CAPES (Bertha API) → {len(resultados)} oportunidades individuais extraídas")
            return resultados

    soup = get_html(url_base)
    if not soup:
        return resultados

    def limpar_texto(texto: str) -> str:
        texto = re.sub(r"\s+", " ", texto or "").strip(" -\n\t")
        texto = re.sub(r"\s+([,.;:])", r"\1", texto)
        return texto

    def link_de_oportunidade(href: str) -> bool:
        link = abs_url(href, url_base)
        if not link:
            return False

        p = urlparse(link)
        host = p.netloc.lower()
        path = p.path.lower()
        query = p.query.lower()

        if any(x in host for x in ("facebook.", "twitter.", "x.com", "linkedin.", "whatsapp.")):
            return False
        if path.endswith((".jpg", ".jpeg", ".png", ".gif", ".svg", ".css", ".js", ".ico")):
            return False

        if host.endswith("bertha.capes.gov.br"):
            return any(x in path for x in ("/oportunidade", "/oportunidades/", "/edital", "/programa")) or "id=" in query

        if host in ("www.gov.br", "gov.br") and path.startswith("/capes/"):
            return any(x in path for x in (
                "/acoes-e-programas/",
                "/bolsas/",
                "/editais",
                "/assuntos/editais",
                "/assuntos/noticias/",
                "/centrais-de-conteudo/editais/",
            ))

        if host.endswith("capes.gov.br"):
            return any(x in path for x in ("/inscricao", "/edital", "/oportunidade", "/programa")) or "id=" in query

        return False

    def container_da_oportunidade(a):
        candidatos = []
        for pai in a.parents:
            if getattr(pai, "name", None) in ("body", "html"):
                break
            texto = limpar_texto(pai.get_text(" ", strip=True))
            links = [x for x in pai.find_all("a", href=True) if link_de_oportunidade(x["href"])]
            if len(texto) >= 20 and len(links) <= 4:
                candidatos.append((len(texto), pai))
            if pai.name in ("article", "li", "tr", "section"):
                break
        return min(candidatos, key=lambda x: x[0])[1] if candidatos else a

    def titulo_da_oportunidade(a, container) -> str:
        texto_link = limpar_texto(a.get_text(" ", strip=True))
        texto_link_baixo = texto_link.lower()
        rotulos_genericos = {
            "acesse", "acessar", "abrir", "detalhes", "saiba mais", "ver mais",
            "inscricoes", "inscrições", "candidatar", "candidatar-se", "link",
        }

        for seletor in ("h1", "h2", "h3", "h4", "h5", "[class*=titulo]", "[class*=title]", "strong"):
            h = container.select_one(seletor)
            if h:
                titulo = limpar_texto(h.get_text(" ", strip=True))
                if len(titulo) >= 10 and titulo.lower() not in rotulos_genericos:
                    return titulo

        if len(texto_link) >= 10 and texto_link_baixo not in rotulos_genericos:
            return texto_link

        textos = []
        for t in container.stripped_strings:
            t = limpar_texto(t)
            if len(t) < 10 or t.lower() in rotulos_genericos:
                continue
            if re.search(r"^(prazo|inscri|publica|saiba|acesse|ver mais)", t, re.I):
                continue
            textos.append(t)

        if textos:
            titulo = textos[0]
            if len(textos) > 1 and re.search(r"\bedital\b", textos[1], re.I) and textos[1] not in titulo:
                titulo = f"{titulo} - {textos[1]}"
            return limpar_texto(titulo)

        return texto_link

    def prazo_da_oportunidade(container) -> str:
        texto = container.get_text(" ", strip=True)
        padroes = [
            r"(?:ate|até|prazo(?:\s+final)?|inscri[cç][oõ]es?.{0,30})(\d{2}/\d{2}/\d{4})",
            r"(\d{2}/\d{2}/\d{4})",
            r"(\d{4}-\d{2}-\d{2})",
        ]
        for padrao in padroes:
            m = re.search(padrao, texto, flags=re.I)
            if m:
                return m.group(1)
        return ""

    links_processados = set()

    for a in soup.find_all("a", href=True):
        if not link_de_oportunidade(a["href"]):
            continue

        link_completo = abs_url(a["href"], url_base)
        if not link_completo or link_completo in links_processados:
            continue

        container = container_da_oportunidade(a)
        titulo = titulo_da_oportunidade(a, container)
        if len(titulo) < 10 or not area_relevante(f"{titulo} {container.get_text(' ', strip=True)}"):
            continue

        resultados.append(bolsa(
            titulo,
            "CAPES (Bertha)",
            "Brasil",
            detectar_nivel(titulo),
            prazo_da_oportunidade(container),
            link_completo,
        ))
        links_processados.add(link_completo)

    log.info(f"CAPES (Bertha) → {len(resultados)} oportunidades individuais extraídas")
    return resultados

def coletar_cnpq() -> list[dict]:
    resultados = []
    urls = [
        "https://www.gov.br/cnpq/pt-br/acesso-a-informacao/acoes-e-programas/programas/bolsas-no-pais",
        "https://www.gov.br/cnpq/pt-br/acesso-a-informacao/acoes-e-programas/programas/bolsas-no-exterior",
        "https://www.gov.br/cnpq/pt-br/assuntos/noticias",  
    ]
    import re
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for item in soup.select("article, .tileItem, .summary, li"):
            h = item.find(["h2", "h3", "h4"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 12 or not area_relevante(titulo):
                continue
            
            link = encontrar_melhor_link(item, url)
            
            prazo = ""
            m = re.search(r"\d{2}/\d{2}/\d{4}", item.get_text())
            if m:
                prazo = m.group()
            resultados.append(bolsa(titulo, "CNPq", "Brasil", detectar_nivel(titulo), prazo, link))

    log.info(f"CNPq → {len(resultados)} editais")
    return resultados

def coletar_cnpq_chamadas_abertas() -> list[dict]:
    """
    Coletor específico para chamadas abertas do CNPq.
    Mantém cada chamada em uma linha própria e ignora menus/PDFs duplicados.
    """
    import re

    resultados = []
    url_base = "https://www.gov.br/cnpq/pt-br/chamadas/abertas-para-submissao"
    soup = get_html(url_base)
    if not soup:
        return resultados

    chamadas_processadas = set()
    for a in soup.find_all("a", href=True):
        titulo = re.sub(r"\s+", " ", a.get_text(" ", strip=True)).strip()
        if not titulo.lower().startswith("chamada") or len(titulo) < 20:
            continue

        link = abs_url(a["href"], url_base)
        if not link:
            continue

        p = urlparse(link)
        path = p.path.lower()
        if p.netloc.lower() not in ("www.gov.br", "gov.br") or not path.startswith("/cnpq/"):
            continue
        if path.endswith((".pdf", ".doc", ".docx", ".xls", ".xlsx")):
            continue
        if "/chamadas/todas-as-chamadas/" not in path:
            continue

        chave = f"{titulo.lower()}::{link.lower()}"
        if chave in chamadas_processadas:
            continue

        container = a.find_parent(["article", "li", "tr", "div", "section"])
        texto_container = container.get_text(" ", strip=True) if container else titulo
        prazo = ""
        periodo = re.search(
            r"Inscri[cç][oõ]es?:\s*(\d{2}/\d{2}/\d{4})\s*(?:a|até|ate|-)\s*(\d{2}/\d{2}/\d{4})",
            texto_container,
            flags=re.I,
        )
        if periodo:
            prazo = periodo.group(2)
        for padrao in (
            r"(?:submiss[aã]o|inscri[cç][aã]o|prazo|até|ate).{0,40}(\d{2}/\d{2}/\d{4})",
            r"(\d{2}/\d{2}/\d{4})",
        ):
            if prazo:
                break
            m = re.search(padrao, texto_container, flags=re.I)
            if m:
                prazo = m.group(1)
                break

        resultados.append(bolsa(
            titulo,
            "CNPq (Chamadas Abertas)",
            "Brasil",
            detectar_nivel(titulo),
            prazo,
            link,
            area="Chamada Publica",
        ))
        chamadas_processadas.add(chave)

    log.info(f"CNPq Chamadas Abertas → {len(resultados)} chamadas individuais")
    return resultados

def coletar_fapesp() -> list[dict]:
    resultados = []
    feed = feedparser.parse("https://www.fapesp.br/rss/oportunidades.xml")
    for e in feed.entries:
        titulo = e.get("title", "").strip()
        if not titulo or not area_relevante(titulo):
            continue
        resultados.append(bolsa(
            titulo, "FAPESP", "Brasil",
            detectar_nivel(titulo),
            e.get("published", "")[:10],
            e.get("link", ""),
        ))
    log.info(f"FAPESP → {len(resultados)} editais via RSS")
    return resultados

def coletar_fapesp_chamadas() -> list[dict]:
    """
    Coletor específico para https://fapesp.br/chamadas/.
    Lista cada chamada aberta em linha individual, preservando o link próprio
    de cada oportunidade e evitando menus, links auxiliares e repetições.
    """
    import re

    resultados = []
    url_base = "https://fapesp.br/chamadas/"
    soup = get_html(url_base)
    if not soup:
        return resultados

    inicio_lista = "/18132"
    fim_lista = "/10273"
    links_auxiliares = {
        "auxílio à pesquisa regular",
        "auxilio a pesquisa regular",
        "projeto temático",
        "projeto tematico",
        "individual research grant",
        "coordinated programmes",
    }
    vistos = set()
    coletando = False

    def texto_limpo(elemento) -> str:
        return re.sub(r"\s+", " ", elemento.get_text(" ", strip=True)).strip()

    for a in soup.find_all("a", href=True):
        titulo = texto_limpo(a)
        link = abs_url(a.get("href", ""), url_base)
        if not titulo or not link:
            continue

        partes = urlparse(link)
        dominio = partes.netloc.lower()
        caminho = partes.path.rstrip("/") or "/"

        if caminho == inicio_lista:
            coletando = True
        if not coletando:
            continue

        eh_chamada_fapesp = (
            "fapesp.br" in dominio
            and (re.fullmatch(r"/\d+", caminho) or caminho == "/acordos-dfg")
        )
        if eh_chamada_fapesp and titulo.lower() not in links_auxiliares:
            chave = link.rstrip("/").lower()
            if chave not in vistos:
                resultados.append(bolsa(
                    titulo,
                    "FAPESP (Chamadas)",
                    "Brasil",
                    detectar_nivel(titulo),
                    "",
                    link,
                    area="Chamada Publica",
                ))
                vistos.add(chave)

        if caminho == fim_lista:
            break

    log.info(f"FAPESP Chamadas → {len(resultados)} chamadas individuais")
    return resultados

def coletar_fappr_programas_abertos() -> list[dict]:
    """
    Coletor específico para Programas Abertos da Fundação Araucária/FAPPR.
    Lista cada Chamada Pública (CP) aberta em linha própria, usando o edital
    como link principal e ignorando anexos, atos e resultados.
    """
    import re

    resultados = []
    url_base = "https://www.fappr.pr.gov.br/Programas-Abertos"
    soup = get_html(url_base)
    if not soup:
        return resultados

    def texto_limpo(elemento) -> str:
        return re.sub(r"\s+", " ", elemento.get_text(" ", strip=True)).strip()

    def pai_com_classe(elemento, classe: str):
        for pai in elemento.parents:
            classes = pai.get("class", []) if hasattr(pai, "get") else []
            if classe in classes:
                return pai
        return None

    def prazo_da_secao(texto: str) -> str:
        m = re.search(
            r"Inscri[cç][oõ]es?:\s*(?:at[eé]\s*)?(\d{2}/\d{2}/\d{4})",
            texto,
            flags=re.I,
        )
        if m:
            return m.group(1)
        if re.search(r"fluxo\s+cont[ií]nuo", texto, flags=re.I):
            return "Fluxo continuo"
        return ""

    chamadas_processadas = set()
    for h in soup.find_all(["h3", "h4"]):
        cabecalho = texto_limpo(h)
        if not re.match(r"^CP\s+\d{2}/\d{2,4}\b", cabecalho, flags=re.I):
            continue

        bloco_principal = pai_com_classe(h, "col-main") or h
        bloco_linha = pai_com_classe(h, "two-col-right") or bloco_principal
        texto_bloco = texto_limpo(bloco_linha)

        chamada_completa = ""
        m = re.search(r"Chamada\s+P[úu]blica\s+\d{2}/\d{4}\s*-\s*([^\"”]+)", texto_bloco, flags=re.I)
        if m:
            chamada_completa = re.sub(r"\s+", " ", m.group(0)).strip()

        titulo = cabecalho
        if chamada_completa and chamada_completa.lower() not in titulo.lower():
            titulo = f"{cabecalho} - {chamada_completa}"

        edital = None
        for a in bloco_principal.find_all("a", href=True):
            rotulo = texto_limpo(a)
            if "edital" in rotulo.lower():
                edital = abs_url(a["href"], url_base)
                break
        if not edital:
            continue

        chave = f"{cabecalho.lower()}::{edital.lower()}"
        if chave in chamadas_processadas:
            continue

        resultados.append(bolsa(
            titulo,
            "FAPPR / Fundacao Araucaria",
            "Brasil",
            detectar_nivel(f"{titulo} {texto_bloco}"),
            prazo_da_secao(texto_bloco),
            edital,
            area="Chamada Publica",
        ))
        chamadas_processadas.add(chave)

    log.info(f"FAPPR Programas Abertos → {len(resultados)} chamadas individuais")
    return resultados

# ── EUROPA ───────────────────────────────────────────────────────────────────

def coletar_ukri() -> list[dict]:
    resultados = []
    termos = ["management", "marketing", "artificial intelligence",
              "entrepreneurship", "economics", "finance", "innovation", "data science"]

    for termo in termos:
        url = (
            f"https://gtr.ukri.org/gtr/api/projects"
            f"?q={termo.replace(' ', '+')}&f=pro.s%3AACTIVE&fetchSize=20"
        )
        data = get_json(url, headers={"Accept": "application/json"})
        if not data:
            continue
        projetos = data.get("project", []) if isinstance(data, dict) else []
        for p in projetos:
            titulo = p.get("title", "").strip()
            if not titulo:
                continue
            link = f"https://gtr.ukri.org/projects?ref={p.get('grantReference','')}"
            resultados.append(bolsa(
                titulo, "UKRI / GTR", "Reino Unido",
                detectar_nivel(titulo),
                p.get("endDate", "")[:10],
                link,
                area=termo.title(),
            ))
        time.sleep(0.5)

    feed = feedparser.parse("https://www.ukri.org/opportunity/feed/")
    for e in feed.entries:
        titulo = e.get("title", "").strip()
        if not titulo or not area_relevante(titulo):
            continue
        resultados.append(bolsa(
            titulo, "UKRI", "Reino Unido",
            detectar_nivel(titulo),
            e.get("published", "")[:10],
            e.get("link", ""),
        ))

    log.info(f"UKRI → {len(resultados)} editais")
    return resultados

def coletar_cordis() -> list[dict]:
    resultados = []
    termos = ["management", "marketing", "artificial intelligence",
              "entrepreneurship", "economics", "innovation"]

    for termo in termos:
        url_json = (
            "https://cordis.europa.eu/api/project/opendata/pl/1"
            f"?q={termo.replace(' ', '+')}&format=json"
        )
        data = get_json(url_json)
        if not data:
            continue
        projetos = data if isinstance(data, list) else data.get("results", [])
        for p in projetos[:10]:
            titulo = p.get("title") or p.get("acronym") or ""
            titulo = titulo.strip()
            if not titulo:
                continue
            link = f"https://cordis.europa.eu/project/id/{p.get('id','')}"
            prazo = p.get("endDate", "")[:10]
            resultados.append(bolsa(titulo, "CORDIS / EU", "Europa",
                                    detectar_nivel(titulo), prazo, link, area=termo.title()))
        time.sleep(0.5)

    log.info(f"CORDIS → {len(resultados)} projetos")
    return resultados

def coletar_msca() -> list[dict]:
    resultados = []
    urls = [
        "https://marie-sklodowska-curie-actions.ec.europa.eu/calls",
        "https://marie-sklodowska-curie-actions.ec.europa.eu/calls?type=doctoral-networks",
        "https://marie-sklodowska-curie-actions.ec.europa.eu/calls?type=postdoctoral-fellowships",
    ]
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for card in soup.select(".call-item, .ecl-card, article, li.call"):
            h = card.find(["h2", "h3", "h4"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 10:
                continue
            
            link = encontrar_melhor_link(card, url)
            
            prazo = ""
            for span in card.find_all(["span", "p", "div"]):
                t = span.get_text(strip=True).lower()
                if "deadline" in t or "closing" in t or "prazo" in t:
                    import re
                    m = re.search(r"\d{1,2}[\s/.-]\w+[\s/.-]\d{2,4}", span.get_text())
                    if m:
                        prazo = m.group()
                        break
            resultados.append(bolsa(titulo, "MSCA", "Europa", detectar_nivel(titulo), prazo, link))

    log.info(f"MSCA → {len(resultados)} editais")
    return resultados

def coletar_eu_funding_tenders() -> list[dict]:
    """
    Coletor específico para a busca filtrada do EU Funding & Tenders Portal.
    A URL monitorada pede a primeira página com 100 oportunidades abertas/futuras.
    """
    import re

    resultados = []
    url_monitorada = (
        "https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/"
        "opportunities/calls-for-proposals?order=DESC&pageNumber=1&pageSize=100"
        "&sortBy=startDate&isExactMatch=true&status=31094501,31094502"
        "&startDate=1767236400000,1785466800000"
        "&deadlineDate=1780196400000,1798686000000"
    )
    api_url = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"

    params = {
        "apiKey": "SEDIA",
        "text": "***",
        "pageSize": "100",
        "pageNumber": "1",
    }
    query = {
        "bool": {
            "must": [
                {"terms": {"type": ["1", "2", "8"]}},
                {"terms": {"status": ["31094501", "31094502"]}},
                {"range": {"startDate": {"gte": 1767236400000, "lte": 1785466800000}}},
                {"range": {"deadlineDate": {"gte": 1780196400000, "lte": 1798686000000}}},
            ]
        }
    }
    sort = {"order": "DESC", "field": "startDate"}
    display_fields = [
        "type", "identifier", "reference", "callccm2Id", "title", "status",
        "caName", "projectAcronym", "startDate", "deadlineDate", "deadlineModel",
        "frameworkProgramme", "typesOfAction", "description", "url",
    ]

    def primeiro(meta: dict, campo: str) -> str:
        valor = meta.get(campo, "")
        if isinstance(valor, list):
            return str(valor[0] if valor else "").strip()
        return str(valor or "").strip()

    def data_iso(valor: str) -> str:
        if not valor:
            return ""
        m = re.match(r"(\d{4}-\d{2}-\d{2})", valor)
        return m.group(1) if m else valor[:10]

    try:
        files = {
            "sort": ("blob", json.dumps(sort), "application/json"),
            "query": ("blob", json.dumps(query), "application/json"),
            "languages": ("blob", json.dumps(["en"]), "application/json"),
            "displayFields": ("blob", json.dumps(display_fields), "application/json"),
        }
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://ec.europa.eu",
            "Referer": url_monitorada,
            "X-Requested-With": "XMLHttpRequest",
        }
        r = SESSION.post(api_url, params=params, files=files, headers=headers, timeout=max(TIMEOUT, 45))
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.warning(f"EU Funding & Tenders erro: {e}")
        return resultados

    vistos = set()
    for item in data.get("results", [])[:100]:
        meta = item.get("metadata", {}) or {}
        titulo = primeiro(meta, "title") or item.get("content") or item.get("summary") or ""
        titulo = re.sub(r"\s+", " ", str(titulo)).strip()
        if not titulo:
            continue

        identificador = primeiro(meta, "identifier")
        link = primeiro(meta, "url") or item.get("url") or url_monitorada
        referencia = str(item.get("reference") or "").strip().lower()
        chave = f"{referencia}::{identificador.lower()}::{titulo.lower()}::{link.lower()}"
        if chave in vistos:
            continue

<<<<<<< HEAD
def coletar_euraxess_it_fr_es():
    """
    Busca contratos, financiamentos e oportunidades de hosting no EURAXESS
    para Italia, Franca e Espanha, mantendo apenas resultados relacionados a
    marketing, administracao/gestao e inteligencia artificial.
    """
    bolsas = []
    url_base = "https://euraxess.ec.europa.eu/jobs/search"
    paises = {"Italia": "781", "Franca": "793", "Espanha": "788"}
    tipos = {
        "job_offer": "Contrato / Vaga",
        "funding": "Financiamento",
        "hosting": "Hosting Offer",
    }
    termos = re.compile(
        r"\b("
        r"artificial intelligence|machine learning|deep learning|data science|"
        r"big data|analytics|algorithm|algorithms|intelig[eê]ncia artificial|"
        r"marketing|consumer|brand|branding|advertising|communication|"
        r"management|business administration|administration|organizational|"
        r"governance|leadership|strategy|innovation|entrepreneurship|"
        r"gest[aã]o|administra[cç][aã]o|inova[cç][aã]o|empreendedorismo"
        r")\b",
        flags=re.I,
    )
    headers = {"User-Agent": "Mozilla/5.0"}
    vistos = set()
    sessao = requests.Session()
    sessao.headers.update(headers)

    def limpar_texto(elemento):
        return re.sub(r"\s+", " ", elemento.get_text(" ", strip=True)).strip()

    def extrair_prazo(texto):
        m = re.search(r"Application Deadline:\s*(.*?)(?:\s+Work Locations:|$)", texto, flags=re.I)
        return m.group(1).strip() if m else "Consultar edital"

    def pagina_filtrada(id_pais, tipo):
        inicial = sessao.get(url_base, timeout=20)
        inicial.raise_for_status()
        soup_inicial = BeautifulSoup(inicial.text, "html.parser")
        form = soup_inicial.find("form", {"id": "oe-list-pages-facets-form"})
        dados = {}
        if form:
            for inp in form.find_all("input"):
                nome = inp.get("name")
                if nome:
                    dados[nome] = inp.get("value", "")
        dados["job_country[]"] = id_pais
        dados["offer_type[]"] = tipo
        resposta = sessao.post(url_base, data=dados, timeout=20)
        resposta.raise_for_status()
        return BeautifulSoup(resposta.text, "html.parser")

    for pais, id_pais in paises.items():
        for tipo, rotulo_tipo in tipos.items():
            try:
                soup = pagina_filtrada(id_pais, tipo)
            except Exception as e:
                log.warning(f"EURAXESS {pais} erro: {e}")
                continue

            for item in soup.find_all("article"):
                a = item.select_one('h3 a[href*="/jobs/"]')
                if not a:
                    continue

                titulo = limpar_texto(a)
                texto = limpar_texto(item)
                termo = termos.search(f"{titulo} {texto}")
                if not titulo or not termo:
                    continue

                link = urljoin("https://euraxess.ec.europa.eu", a.get("href", ""))
                chave = f"{pais.lower()}::{tipo}::{link.rstrip('/').lower()}"
                if chave in vistos:
                    continue

                bolsas.append({
                    "titulo": titulo,
                    "fonte": "EURAXESS",
                    "pais": pais,
                    "area": f"{rotulo_tipo} - {termo.group(1).title()}",
                    "nivel": detectar_nivel(f"{titulo.lower()} {texto.lower()}"),
                    "prazo": extrair_prazo(texto),
                    "link": link,
                    "data_coleta": str(date.today()),
                    "ativa": "Sim"
                })
                vistos.add(chave)

    log.info(f"EURAXESS Italia/Franca/Espanha: {len(bolsas)} oportunidades filtradas")
    return bolsas


def coletar_bolsas():
    bolsas = []
=======
        tipo_acao = primeiro(meta, "typesOfAction")
        programa = primeiro(meta, "frameworkProgramme")
        area = tipo_acao or "EU Funding Call"
        if programa and tipo_acao:
            area = f"{tipo_acao} / {programa}"
>>>>>>> f9770e4 (Atualiza pagina de bolsas no GitHub Pages)

        titulo_final = f"{identificador} - {titulo}" if identificador and identificador not in titulo else titulo
        resultados.append(bolsa(
            titulo_final,
            "EU Funding & Tenders",
            "Europa",
            detectar_nivel(f"{titulo_final} {tipo_acao}"),
            data_iso(primeiro(meta, "deadlineDate")),
            link,
            area=area,
        ))
        vistos.add(chave)

    total = data.get("totalResults", len(resultados))
    log.info(f"EU Funding & Tenders → {len(resultados)} oportunidades da página 1 | {total} resultados na busca")
    return resultados

def coletar_daad() -> list[dict]:
    resultados = []
    urls = [
        "https://www.daad-brasil.org/pt/encontrar-bolsas-e-programas/programas-daad/",
        "https://www.daad-brasil.org/pt/encontrar-bolsas-e-programas/financiamentos-para-brasileiros/",
    ]
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for card in soup.select(".program-item, .scholarship-item, article, .entry, li.programa"):
            h = card.find(["h2", "h3", "h4"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 10:
                continue
            
            link = encontrar_melhor_link(card, url)
            
            resultados.append(bolsa(titulo, "DAAD", "Alemanha", detectar_nivel(titulo), "", link))

    log.info(f"DAAD → {len(resultados)} programas")
    return resultados

def coletar_erasmus() -> list[dict]:
    resultados = []
    url = "https://erasmus-plus.ec.europa.eu/opportunities/individuals"
    soup = get_html(url)
    if not soup:
        return resultados
    for card in soup.select(".ecl-card, .opportunity-item, article"):
        h = card.find(["h2", "h3", "h4"])
        if not h:
            continue
        titulo = h.get_text(strip=True)
        if len(titulo) < 10 or not area_relevante(titulo):
            continue
        
        link = encontrar_melhor_link(card, url)
        
        resultados.append(bolsa(titulo, "Erasmus+", "Europa", detectar_nivel(titulo), "", link))
    log.info(f"Erasmus+ → {len(resultados)} oportunidades")
    return resultados

# ── AMERICA DO NORTE ──────────────────────────────────────────────────────────

def coletar_fulbright() -> list[dict]:
    resultados = []
    urls = [
        "https://fulbrightscholars.org/grants#all",
        "https://fulbright.org.br/bolsas/",
        "https://fulbright.org.br/programas/",
    ]
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for card in soup.select(".grant-item, .views-row, article, .program-card, li.bolsa"):
            h = card.find(["h2", "h3", "h4", "a"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 10:
                continue
            
            link = encontrar_melhor_link(card, url)
            
            import re
            prazo = ""
            m = re.search(r"\w+ \d{1,2},? \d{4}", card.get_text())
            if m:
                prazo = m.group()
            resultados.append(bolsa(titulo, "Fulbright", "EUA", detectar_nivel(titulo), prazo, link))

    log.info(f"Fulbright → {len(resultados)} grants")
    return resultados

<<<<<<< HEAD
    bolsas += coletar_euraxess_it_fr_es()

    # Horizon (manual — melhor abordagem)
    bolsas.append({
        "titulo": "Horizon Europe – Funding & Tenders Portal",
        "fonte": "Horizon Europe",
        "pais": "Europa",
        "area": "Pesquisa",
        "nivel": "Doutorado/Pós-doc",
        "prazo": "Variável",
        "link": "https://ec.europa.eu/info/funding-tenders/opportunities/portal/",
        "data_coleta": str(date.today()),
        "ativa": "Sim"
    })
=======
def coletar_nih_reporter() -> list[dict]:
    resultados = []
    url = "https://api.reporter.nih.gov/v2/projects/search"
    termos = ["management", "marketing", "artificial intelligence",
              "health economics", "entrepreneurship", "innovation"]
    for termo in termos:
        payload = {
            "criteria": {
                "project_titles": [termo],
                "project_status": "active",
            },
            "offset": 0,
            "limit": 15,
            "sort_field": "project_start_date",
            "sort_order": "desc",
        }
        try:
            r = SESSION.post(url, json=payload, timeout=TIMEOUT)
            data = r.json()
            for p in data.get("results", []):
                titulo = p.get("project_title", "").strip()
                if not titulo or not area_relevante(titulo):
                    continue
                link = f"https://reporter.nih.gov/project-details/{p.get('appl_id','')}"
                prazo = p.get("project_end_date", "")[:10]
                resultados.append(bolsa(titulo, "NIH Reporter", "EUA",
                                        detectar_nivel(titulo), prazo, link, area=termo.title()))
            time.sleep(0.5)
        except Exception as e:
            log.warning(f"NIH [{termo}]: {e}")
>>>>>>> f9770e4 (Atualiza pagina de bolsas no GitHub Pages)

    log.info(f"NIH Reporter → {len(resultados)} projetos")
    return resultados

# ── PORTUGAL ─────────────────────────────────────────────────────────────────

def coletar_fct() -> list[dict]:
    resultados = []
    urls = [
        "https://www.fct.pt/apoios/bolsas/candidaturas/",
        "https://www.fct.pt/financiamento/bolsas-de-investigacao/",
    ]
    import re
    for url in urls:
        soup = get_html(url)
        if not soup:
            continue
        for item in soup.select("article, .bolsa-item, li, .entry"):
            h = item.find(["h2", "h3", "h4", "a"])
            if not h:
                continue
            titulo = h.get_text(strip=True)
            if len(titulo) < 12:
                continue
            
            link = encontrar_melhor_link(item, url)
            
            prazo = ""
            m = re.search(r"\d{2}/\d{2}/\d{4}", item.get_text())
            if m:
                prazo = m.group()
            resultados.append(bolsa(titulo, "FCT Portugal", "Portugal",
                                    detectar_nivel(titulo), prazo, link))

    log.info(f"FCT Portugal → {len(resultados)} bolsas")
    return resultados

# ── GLOBAL / AGREGADORES ─────────────────────────────────────────────────────

def coletar_openaire() -> list[dict]:
    resultados = []
    termos = ["management", "marketing", "artificial intelligence",
              "entrepreneurship", "innovation", "economics"]
    for termo in termos:
        url = (
            f"https://api.openaire.eu/search/projects"
            f"?keywords={termo.replace(' ', '+')}&size=10"
            f"&format=json&funder=ec__________::EC"
        )
        data = get_json(url)
        if not data:
            continue
        try:
            projetos = (data.get("response", {})
                           .get("results", {})
                           .get("result", []))
            for p in projetos:
                meta = p.get("metadata", {}).get("oaf:entity", {}).get("oaf:project", {})
                titulo = meta.get("title", {})
                titulo = titulo.get("$", "") if isinstance(titulo, dict) else str(titulo)
                titulo = titulo.strip()
                if not titulo or len(titulo) < 10:
                    continue
                link = meta.get("websiteurl", {})
                link = link.get("$", "") if isinstance(link, dict) else ""
                prazo = meta.get("enddate", {})
                prazo = prazo.get("$", "")[:10] if isinstance(prazo, dict) else ""
                resultados.append(bolsa(titulo, "OpenAire / EU", "Europa",
                                        detectar_nivel(titulo), prazo, link or "https://explore.openaire.eu",
                                        area=termo.title()))
        except Exception as e:
            log.warning(f"OpenAire parse [{termo}]: {e}")
        time.sleep(0.4)

    log.info(f"OpenAire → {len(resultados)} projetos")
    return resultados

def coletar_findaphd() -> list[dict]:
    resultados = []
    feeds = [
        ("https://www.findaphd.com/phds/rss.aspx?Keywords=marketing&MastersType=0",          "Marketing"),
        ("https://www.findaphd.com/phds/rss.aspx?Keywords=management&MastersType=0",         "Management"),
        ("https://www.findaphd.com/phds/rss.aspx?Keywords=artificial+intelligence",           "IA"),
        ("https://www.findaphd.com/phds/rss.aspx?Keywords=entrepreneurship",                  "Empreendedorismo"),
        ("https://www.findaphd.com/phds/rss.aspx?Keywords=economics+finance",                "Economia/Financas"),
    ]
    for feed_url, area in feeds:
        feed = feedparser.parse(feed_url)
        for e in feed.entries[:15]:
            titulo = e.get("title", "").strip()
            if not titulo:
                continue
            resultados.append(bolsa(
                titulo, "FindAPhD", "Internacional",
                "Doutorado", e.get("published", "")[:10],
                e.get("link", ""), area=area,
            ))
        time.sleep(0.3)

    log.info(f"FindAPhD → {len(resultados)} vagas de PhD")
    return resultados

def coletar_scholarshipportal() -> list[dict]:
    resultados = []
    feeds = [
        "https://www.scholarshipportal.com/rss",
        "https://www.scholarshipportal.com/rss?subject=business-administration",
        "https://www.scholarshipportal.com/rss?subject=economics",
    ]
    for feed_url in feeds:
        feed = feedparser.parse(feed_url)
        for e in feed.entries[:20]:
            titulo = e.get("title", "").strip()
            if not titulo or not area_relevante(titulo):
                continue
            resultados.append(bolsa(
                titulo, "ScholarshipPortal", "Internacional",
                detectar_nivel(titulo), e.get("published", "")[:10],
                e.get("link", ""),
            ))
    log.info(f"ScholarshipPortal → {len(resultados)} bolsas")
    return resultados

# ═════════════════════════════════════════════════════════════════════════════
# ORQUESTRADOR E DEDUPLICAÇÃO
# ═════════════════════════════════════════════════════════════════════════════

COLETORES = [
    ("CAPES",            coletar_capes),
    ("CAPES (Bertha)",   coletar_capes_bertha), # NOVO COLETOR ADICIONADO
    ("CNPq",             coletar_cnpq),
    ("CNPq Chamadas Abertas", coletar_cnpq_chamadas_abertas),
    ("FAPESP",           coletar_fapesp),
    ("FAPESP Chamadas",  coletar_fapesp_chamadas),
    ("FAPPR Programas Abertos", coletar_fappr_programas_abertos),
    ("UKRI",             coletar_ukri),
    ("CORDIS",           coletar_cordis),
    ("MSCA",             coletar_msca),
    ("EU Funding & Tenders", coletar_eu_funding_tenders),
    ("DAAD",             coletar_daad),
    ("Erasmus+",         coletar_erasmus),
    ("OpenAire",         coletar_openaire),
    ("Fulbright",        coletar_fulbright),
    ("NIH Reporter",     coletar_nih_reporter),
    ("FCT Portugal",     coletar_fct),
    ("FindAPhD",         coletar_findaphd),
    ("ScholarshipPortal",coletar_scholarshipportal),
]

def coletar_tudo() -> list[dict]:
    todas = []
    for nome, func in COLETORES:
        try:
            log.info(f"Iniciando: {nome}...")
            parcial = func()
            todas.extend(parcial)
        except Exception as e:
            log.error(f"{nome} — erro inesperado: {e}")
        time.sleep(PAUSA_ENTRE_FONTES)

    vistos_agora = set()
    unicas = []
    for b in todas:
        chave = f"{b['fonte']}::{b['titulo'][:80].lower()}::{b.get('link', '').lower()}"
        if chave not in vistos_agora:
            vistos_agora.add(chave)
            unicas.append(b)

    log.info(f"Total coletado: {len(todas)} → {len(unicas)} após deduplicação interna")
    return unicas

def carregar_ids() -> set:
    if os.path.exists(ARQUIVO_IDS):
        try:
            with open(ARQUIVO_IDS, encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()

def salvar_ids(ids: set):
    with open(ARQUIVO_IDS, "w", encoding="utf-8") as f:
        json.dump(sorted(ids), f, ensure_ascii=False)

def gerar_id(fonte: str, titulo: str, link: str = "") -> str:
    chave = f"{fonte}::{titulo.strip().lower()[:120]}::{(link or '').strip().lower()}"
    return hashlib.md5(chave.encode()).hexdigest()

def gerar_id_legacy(fonte: str, titulo: str) -> str:
    chave = f"{fonte}::{titulo.strip().lower()[:120]}"
    return hashlib.md5(chave.encode()).hexdigest()

# ═════════════════════════════════════════════════════════════════════════════
# EXCEL, HTML E GIT
# ═════════════════════════════════════════════════════════════════════════════

CABECALHOS = ["Título", "Fonte", "País", "Área", "Nível", "Prazo", "Link", "Data Coleta"]
COLUNAS    = ["titulo", "fonte", "pais", "area", "nivel", "prazo", "link", "data"]
COR_HDR    = "1F3864"
COR_PAR    = "EBF0FA"
COR_NOVA   = "FFF2CC"

def _cabecalho(ws):
    for i, cab in enumerate(CABECALHOS, 1):
        c = ws.cell(row=1, column=i, value=cab)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=COR_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([70, 18, 16, 24, 18, 13, 60, 13], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def salvar_excel(bolsas: list[dict], ids_vistos: set) -> int:
    if os.path.exists(ARQUIVO_EXCEL):
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bolsas"
        _cabecalho(ws)

    adicionadas = 0
    hoje = str(date.today())

    for b in bolsas:
        if not b.get("titulo"):
            continue
        bid_antigo = gerar_id_legacy(b["fonte"], b["titulo"])
        bid = gerar_id(b["fonte"], b["titulo"], b.get("link", ""))
        if bid in ids_vistos or bid_antigo in ids_vistos:
            continue

        linha = ws.max_row + 1
        for col, campo in enumerate(COLUNAS, 1):
            val = b.get(campo, "")
            c = ws.cell(row=linha, column=col, value=val)
            cor = COR_NOVA if b.get("data") == hoje else (COR_PAR if linha % 2 == 0 else "FFFFFF")
            c.fill = PatternFill("solid", fgColor=cor)
            if col == 7 and val:  
                c.hyperlink = str(val)
                c.font = Font(color="0070C0", underline="single")

        ids_vistos.add(bid)
        ids_vistos.add(bid_antigo)
        adicionadas += 1

    ws.auto_filter.ref = ws.dimensions
    wb.save(ARQUIVO_EXCEL)
    log.info(f"Excel → {adicionadas} novas | {ws.max_row - 1} total acumulado")
    return adicionadas

def gerar_html():
    if not os.path.exists(ARQUIVO_EXCEL):
        log.error("Excel não encontrado. Rode a coleta primeiro.")
        return

    wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
    ws = wb.active
    fontes, paises, niveis, areas = set(), set(), set(), set()
    registros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        r = {
            "titulo": str(row[0] or "").strip(),
            "fonte":  str(row[1] or "").strip(),
            "pais":   str(row[2] or "").strip(),
            "area":   str(row[3] or "").strip(),
            "nivel":  str(row[4] or "").strip(),
            "prazo":  str(row[5] or "").strip(),
            "link":   str(row[6] or "").strip(),
            "data":   str(row[7] or "").strip(),
        }
        if not r["titulo"]:
            continue
        fontes.add(r["fonte"])
        paises.add(r["pais"])
        niveis.add(r["nivel"])
        areas.add(r["area"])
        registros.append(r)

    if not registros:
        return

    def opts(vals: set) -> str:
        return "".join(
            f'<option value="{html_lib.escape(v)}">{html_lib.escape(v)}</option>'
            for v in sorted(vals) if v
        )

    dados_js = json.dumps(registros, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Radar de Bolsas — Administração · Marketing · IA · Inovação</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;background:#f0f2f8;padding:16px;color:#1a1a2e}}
header{{background:#1F3864;color:#fff;border-radius:10px;padding:16px 22px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}}
header h1{{font-size:18px;font-weight:700}}
header p{{font-size:12px;opacity:.75}}
.filtros{{background:#fff;border:0.5px solid #d5daea;border-radius:10px;padding:14px 18px;margin-bottom:12px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}}
.fg{{display:flex;flex-direction:column;gap:3px}}
.fg label{{font-size:11px;color:#666;font-weight:500}}
.fg input,.fg select{{padding:6px 9px;border:1px solid #cdd5e0;border-radius:6px;font-size:12px;color:#222;background:#fff;min-width:140px}}
.fg input{{min-width:200px}}
.resumo{{font-size:12px;color:#555;margin-bottom:8px}}
.resumo strong{{color:#1F3864;font-weight:700}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;font-size:12px;box-shadow:0 1px 6px rgba(0,0,0,.07)}}
th{{background:#1F3864;color:#fff;padding:9px 11px;text-align:left;white-space:nowrap;font-size:11px;letter-spacing:.02em}}
td{{padding:7px 11px;border-bottom:1px solid #eef0f7;vertical-align:top;line-height:1.4}}
td:first-child{{max-width:420px;word-break:break-word;font-weight:500}}
tr.par td{{background:#f7f9ff}}
tr:hover td{{background:#edf1fb!important}}
.badge{{display:inline-flex;align-items:center;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700;white-space:nowrap}}
.b-br{{background:#e8f5e9;color:#1b5e20}}
.b-eu{{background:#e3f2fd;color:#0d47a1}}
.b-us{{background:#fce4ec;color:#880e4f}}
.b-pt{{background:#fff3e0;color:#e65100}}
.b-gl{{background:#f3e5f5;color:#4a148c}}
.b-nivel{{background:#e8edf8;color:#1F3864;margin-left:3px}}
a{{color:#1F3864;text-decoration:none;font-weight:600}}
a:hover{{text-decoration:underline}}
.vazio{{text-align:center;padding:48px;color:#999;font-size:13px}}
.prazo-vazio{{color:#bbb;font-style:italic}}
@media(max-width:700px){{th:nth-child(3),td:nth-child(3),th:nth-child(4),td:nth-child(4){{display:none}}}}
</style>
</head>
<body>

<header>
  <div>
    <h1>Radar de Bolsas de Pesquisa</h1>
    <p>Administração · Marketing · IA · Empreendedorismo · Economia</p>
  </div>
  <p>Atualizado em {date.today().strftime("%d/%m/%Y")}</p>
</header>

<div class="filtros">
  <div class="fg"><label>Buscar no título</label>
    <input id="busca" type="text" placeholder="Ex: marketing, phd, innovation..."></div>
  <div class="fg"><label>Fonte</label>
    <select id="f-fonte"><option value="">Todas</option>{opts(fontes)}</select></div>
  <div class="fg"><label>País / Região</label>
    <select id="f-pais"><option value="">Todos</option>{opts(paises)}</select></div>
  <div class="fg"><label>Área</label>
    <select id="f-area"><option value="">Todas</option>{opts(areas)}</select></div>
  <div class="fg"><label>Nível</label>
    <select id="f-nivel"><option value="">Todos</option>{opts(niveis)}</select></div>
</div>

<p class="resumo">Mostrando <strong id="cnt">0</strong> de <strong>{len(registros)}</strong> bolsas</p>

<table>
  <thead><tr>
    <th>Título do Edital</th>
    <th>Fonte</th>
    <th>País</th>
    <th>Área</th>
    <th>Nível</th>
    <th>Prazo</th>
    <th>Link</th>
  </tr></thead>
  <tbody id="tbody"></tbody>
</table>

<script>
var D={dados_js};
var PAIS_COR={{"Brasil":"b-br","Europa":"b-eu","EUA":"b-us","Reino Unido":"b-eu",
  "Portugal":"b-pt","Alemanha":"b-eu","Internacional":"b-gl"}};

function esc(s){{
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

function cor(pais){{
  for(var k in PAIS_COR){{ if(pais.indexOf(k)>=0) return PAIS_COR[k]; }}
  return 'b-gl';
}}

function render(lista){{
  var tb=document.getElementById('tbody');
  document.getElementById('cnt').textContent=lista.length;
  if(!lista.length){{
    tb.innerHTML='<tr><td colspan="7" class="vazio">Nenhuma bolsa para os filtros selecionados.</td></tr>';
    return;
  }}
  tb.innerHTML=lista.map(function(b,i){{
    var pc=cor(b.pais);
    var pr=b.prazo?esc(b.prazo):'<span class="prazo-vazio">—</span>';
    var lk=b.link?'<a href="'+esc(b.link)+'" target="_blank" rel="noopener">Abrir &rarr;</a>':'—';
    return '<tr class="'+(i%2?'':'par')+'">'
      +'<td>'+esc(b.titulo)+'</td>'
      +'<td><span class="badge '+pc+'">'+esc(b.fonte)+'</span></td>'
      +'<td>'+esc(b.pais)+'</td>'
      +'<td>'+esc(b.area)+'</td>'
      +'<td><span class="badge b-nivel">'+esc(b.nivel)+'</span></td>'
      +'<td>'+pr+'</td>'
      +'<td>'+lk+'</td>'
      +'</tr>';
  }}).join('');
}}

function filtrar(){{
  var busca=document.getElementById('busca').value.toLowerCase();
  var fonte=document.getElementById('f-fonte').value;
  var pais =document.getElementById('f-pais').value;
  var area =document.getElementById('f-area').value;
  var nivel=document.getElementById('f-nivel').value;
  render(D.filter(function(b){{
    if(busca && b.titulo.toLowerCase().indexOf(busca)<0) return false;
    if(fonte && b.fonte!==fonte) return false;
    if(pais  && b.pais!==pais)  return false;
    if(area  && b.area!==area)  return false;
    if(nivel && b.nivel!==nivel) return false;
    return true;
  }}));
}}

['busca','f-fonte','f-pais','f-area','f-nivel'].forEach(function(id){{
  var el=document.getElementById(id);
  el.addEventListener('input',filtrar);
  el.addEventListener('change',filtrar);
}});

render(D);
</script>
</body>
</html>"""

    with open(ARQUIVO_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"HTML gerado com {len(registros)} bolsas → {ARQUIVO_HTML}")

    os.makedirs(PASTA_GITHUB_PAGES, exist_ok=True)
    with open(ARQUIVO_HTML_PAGES, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"HTML do GitHub Pages gerado → {ARQUIVO_HTML_PAGES}")

def publicar():
    git = ["git", "-c", f"safe.directory={PASTA_SAIDA}", "-C", PASTA_SAIDA]
    try:
        os.makedirs(PASTA_GITHUB_PAGES, exist_ok=True)
        shutil.copy2(ARQUIVO_EXCEL, ARQUIVO_EXCEL_PAGES)
        shutil.copy2(ARQUIVO_IDS, ARQUIVO_IDS_PAGES)

        pull = subprocess.run(git + ["pull", "origin", "main", "--rebase", "--autostash"],
                              capture_output=True, text=True)
        if pull.returncode != 0:
            log.warning(f"Git pull avisou: {pull.stderr or pull.stdout}")

        subprocess.run(git + ["add",
                        "index.html", "bolsas_pesquisa.xlsx", "ids_vistos.json",
                        "bolsas/index.html", "bolsas/bolsas_pesquisa.xlsx", "bolsas/ids_vistos.json"],
                       check=True, capture_output=True)
        r = subprocess.run(git + ["commit", "-m",
                             f"update {date.today()}"],
                            capture_output=True, text=True)
        msg_commit = f"{r.stdout}\n{r.stderr}"
        if "nothing to commit" in msg_commit:
            log.info("Git: sem alterações para publicar.")
            return
        if r.returncode != 0:
            raise subprocess.CalledProcessError(
                r.returncode, r.args, output=r.stdout, stderr=r.stderr
            )

        subprocess.run(git + ["push", "origin", "HEAD:main"],
                       check=True, capture_output=True)
        log.info("GitHub Pages atualizado.")
    except subprocess.CalledProcessError as e:
        log.warning(f"Git erro: {e.stderr or e}")

def main():
    log.info("=" * 60)
    log.info("Inicio da coleta — v5 (Com Módulo Bertha)")
    log.info("=" * 60)
    ids_vistos = carregar_ids()
    bolsas     = coletar_tudo()
    novas      = salvar_excel(bolsas, ids_vistos)
    salvar_ids(ids_vistos)
    gerar_html()
    publicar()
    log.info(f"Concluido — {novas} novas bolsas adicionadas.")
    log.info("=" * 60)

if __name__ == "__main__":
    main()
