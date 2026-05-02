"""
Monitor de Bolsas de Pesquisa
==============================
Coleta oportunidades de: CNPq, CAPES, FAPESP, Fulbright, DAAD e feeds RSS.
Salva em Excel no OneDrive (ou pasta local).

Dependências:
    pip install requests beautifulsoup4 openpyxl feedparser lxml

Agendamento: Windows Task Scheduler (ver instrucoes em LEIAME.txt)
"""

import requests
from bs4 import BeautifulSoup
import feedparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import os
import json
import logging
import time

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

# Pasta onde o Excel será salvo
# Troque para o caminho do seu OneDrive se quiser sincronizar automaticamente
# Exemplo: r"C:\Users\SeuNome\OneDrive\Bolsas"
PASTA_SAIDA = os.path.join(os.path.expanduser("~"), "Documents", "BolsasPesquisa")

ARQUIVO_EXCEL = os.path.join(PASTA_SAIDA, "bolsas_pesquisa.xlsx")
ARQUIVO_LOG   = os.path.join(PASTA_SAIDA, "coleta.log")
ARQUIVO_IDS   = os.path.join(PASTA_SAIDA, "ids_vistos.json")  # evita duplicatas

# Timeout para requisições HTTP (segundos)
TIMEOUT = 15

# ─── LOGGING ─────────────────────────────────────────────────────────────────

os.makedirs(PASTA_SAIDA, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(ARQUIVO_LOG, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ─── HEADERS HTTP ────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

# ─── FUNÇÕES AUXILIARES ──────────────────────────────────────────────────────

def carregar_ids_vistos() -> set:
    """Carrega os IDs de bolsas já coletadas para evitar duplicatas."""
    if os.path.exists(ARQUIVO_IDS):
        with open(ARQUIVO_IDS, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def salvar_ids_vistos(ids: set):
    with open(ARQUIVO_IDS, "w", encoding="utf-8") as f:
        json.dump(list(ids), f, ensure_ascii=False, indent=2)


def gerar_id(fonte: str, titulo: str) -> str:
    """ID simples baseado em fonte + título normalizado."""
    return f"{fonte}::{titulo[:80].lower().strip()}"


def get_html(url: str) -> BeautifulSoup | None:
    """Faz GET e retorna BeautifulSoup ou None se falhar."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return BeautifulSoup(r.text, "lxml")
    except Exception as e:
        log.warning(f"Erro ao acessar {url}: {e}")
        return None


# ─── COLETORES ───────────────────────────────────────────────────────────────

def coletar_cnpq() -> list[dict]:
    """
    CNPq — Chamadas abertas.
    URL: https://www.gov.br/cnpq/pt-br/acesso-a-informacao/acoes-e-programas/programas/chamadas-abertas
    """
    bolsas = []
    url = "https://www.gov.br/cnpq/pt-br/acesso-a-informacao/acoes-e-programas/programas/chamadas-abertas"
    soup = get_html(url)
    if not soup:
        return bolsas

    # O CNPq lista chamadas em <article> ou <li> com links
    for item in soup.select("article, li.chamada, div.chamada-item"):
        titulo_tag = item.find(["h2", "h3", "h4", "a"])
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        link_tag = item.find("a", href=True)
        link = link_tag["href"] if link_tag else url
        if link.startswith("/"):
            link = "https://www.gov.br" + link

        # Tenta encontrar prazo
        prazo = ""
        for t in item.find_all(string=True):
            if "prazo" in t.lower() or "encerramento" in t.lower():
                prazo = t.strip()[:60]
                break

        if titulo:
            bolsas.append({
                "titulo": titulo,
                "fonte": "CNPq",
                "pais": "Brasil",
                "area": "Múltiplas",
                "nivel": "Múltiplos",
                "prazo": prazo,
                "link": link,
                "data_coleta": str(date.today()),
                "ativa": "Sim",
            })

    log.info(f"CNPq: {len(bolsas)} bolsas encontradas")
    return bolsas


def coletar_capes() -> list[dict]:
    """
    CAPES — Editais abertos.
    URL: https://www.gov.br/capes/pt-br/acesso-a-informacao/acoes-e-programas/bolsas/bolsas-e-auxilios-internacionais
    """
    bolsas = []
    url = "https://www.gov.br/capes/pt-br/acesso-a-informacao/acoes-e-programas/bolsas"
    soup = get_html(url)
    if not soup:
        return bolsas

    for item in soup.select("article, div.tile, li.programa"):
        titulo_tag = item.find(["h2", "h3", "h4", "a"])
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        link_tag = item.find("a", href=True)
        link = link_tag["href"] if link_tag else url
        if link.startswith("/"):
            link = "https://www.gov.br" + link

        if titulo and len(titulo) > 10:
            bolsas.append({
                "titulo": titulo,
                "fonte": "CAPES",
                "pais": "Brasil",
                "area": "Múltiplas",
                "nivel": "Pós-graduação",
                "prazo": "",
                "link": link,
                "data_coleta": str(date.today()),
                "ativa": "Sim",
            })

    log.info(f"CAPES: {len(bolsas)} bolsas encontradas")
    return bolsas


def coletar_fapesp() -> list[dict]:
    """
    FAPESP — Oportunidades de bolsas.
    Usa o feed RSS oficial: https://www.fapesp.br/rss/oportunidades.xml
    """
    bolsas = []
    feed_url = "https://www.fapesp.br/rss/oportunidades.xml"
    try:
        feed = feedparser.parse(feed_url)
        for entry in feed.entries:
            bolsas.append({
                "titulo": entry.get("title", ""),
                "fonte": "FAPESP",
                "pais": "Brasil",
                "area": "Múltiplas",
                "nivel": "Múltiplos",
                "prazo": entry.get("published", "")[:10],
                "link": entry.get("link", ""),
                "data_coleta": str(date.today()),
                "ativa": "Sim",
            })
    except Exception as e:
        log.warning(f"FAPESP RSS: {e}")

    log.info(f"FAPESP: {len(bolsas)} bolsas encontradas")
    return bolsas


def coletar_fulbright() -> list[dict]:
    """
    Fulbright Brasil — Programas disponíveis.
    URL: https://fulbright.org.br/programas/
    """
    bolsas = []
    url = "https://fulbright.org.br/programas/"
    soup = get_html(url)
    if not soup:
        return bolsas

    for item in soup.select("article, .programa-card, div.entry-summary"):
        titulo_tag = item.find(["h2", "h3", "h4"])
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        link_tag = item.find("a", href=True)
        link = link_tag["href"] if link_tag else url

        nivel = "Múltiplos"
        titulo_lower = titulo.lower()
        if "doctoral" in titulo_lower or "doutorado" in titulo_lower:
            nivel = "Doutorado"
        elif "master" in titulo_lower or "mestrado" in titulo_lower:
            nivel = "Mestrado"
        elif "pesquisador" in titulo_lower or "researcher" in titulo_lower:
            nivel = "Pesquisador"

        if titulo:
            bolsas.append({
                "titulo": titulo,
                "fonte": "Fulbright",
                "pais": "Estados Unidos",
                "area": "Múltiplas",
                "nivel": nivel,
                "prazo": "",
                "link": link,
                "data_coleta": str(date.today()),
                "ativa": "Sim",
            })

    log.info(f"Fulbright: {len(bolsas)} bolsas encontradas")
    return bolsas


def coletar_daad() -> list[dict]:
    """
    DAAD Brasil — Programas de bolsas.
    URL: https://www.daad.org.br/pt/encontrar-bolsas-e-programas/
    """
    bolsas = []
    url = "https://www.daad.org.br/pt/encontrar-bolsas-e-programas/"
    soup = get_html(url)
    if not soup:
        return bolsas

    for item in soup.select(".scholarship-card, article, .program-item, li.bolsa"):
        titulo_tag = item.find(["h2", "h3", "h4", "a"])
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        link_tag = item.find("a", href=True)
        link = link_tag["href"] if link_tag else url
        if link.startswith("/"):
            link = "https://www.daad.org.br" + link

        if titulo and len(titulo) > 8:
            bolsas.append({
                "titulo": titulo,
                "fonte": "DAAD",
                "pais": "Alemanha",
                "area": "Múltiplas",
                "nivel": "Múltiplos",
                "prazo": "",
                "link": link,
                "data_coleta": str(date.today()),
                "ativa": "Sim",
            })

    log.info(f"DAAD: {len(bolsas)} bolsas encontradas")
    return bolsas


def coletar_rss_genericos() -> list[dict]:
    """
    Feeds RSS de portais de bolsas internacionais.
    Adicione mais URLs conforme necessário.
    """
    feeds = [
        {
            "url": "https://www.scholarshipportal.com/rss",
            "fonte": "ScholarshipPortal",
            "pais": "Internacional",
        },
        {
            "url": "https://www.opportunitiesforafricans.com/feed/",
            "fonte": "OpportunitiesPortal",
            "pais": "Internacional",
        },
        # Adicione outros feeds aqui:
        # {"url": "https://exemplo.com/rss", "fonte": "NomeFonte", "pais": "País"},
    ]

    bolsas = []
    for feed_info in feeds:
        try:
            feed = feedparser.parse(feed_info["url"])
            for entry in feed.entries[:20]:  # limita 20 por feed
                bolsas.append({
                    "titulo": entry.get("title", ""),
                    "fonte": feed_info["fonte"],
                    "pais": feed_info["pais"],
                    "area": "Múltiplas",
                    "nivel": "Múltiplos",
                    "prazo": entry.get("published", "")[:10],
                    "link": entry.get("link", ""),
                    "data_coleta": str(date.today()),
                    "ativa": "Sim",
                })
            log.info(f"{feed_info['fonte']}: {len(feed.entries)} entradas no feed")
        except Exception as e:
            log.warning(f"RSS {feed_info['fonte']}: {e}")

    return bolsas


# ─── EXCEL ───────────────────────────────────────────────────────────────────

COLUNAS = ["titulo", "fonte", "pais", "area", "nivel", "prazo", "link", "data_coleta", "ativa"]
CABECALHOS = ["Título", "Fonte", "País", "Área", "Nível", "Prazo", "Link", "Data Coleta", "Ativa"]

COR_CABECALHO = "1F3864"   # azul escuro
COR_LINHA_PAR = "EBF0FA"   # azul bem claro
COR_NOVA      = "FFF2CC"   # amarelo para bolsas novas


def criar_ou_abrir_excel() -> openpyxl.Workbook:
    if os.path.exists(ARQUIVO_EXCEL):
        return openpyxl.load_workbook(ARQUIVO_EXCEL)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bolsas"
    _formatar_cabecalho(ws)
    return wb


def _formatar_cabecalho(ws):
    for col_idx, cab in enumerate(CABECALHOS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Larguras de coluna
    larguras = [60, 14, 16, 18, 16, 14, 50, 13, 7]
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def salvar_excel(novas_bolsas: list[dict], ids_vistos: set) -> int:
    wb = criar_ou_abrir_excel()
    ws = wb["Bolsas"]

    adicionadas = 0
    for bolsa in novas_bolsas:
        bid = gerar_id(bolsa["fonte"], bolsa["titulo"])
        if bid in ids_vistos or not bolsa["titulo"]:
            continue

        linha = ws.max_row + 1
        for col_idx, col in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=linha, column=col_idx, value=bolsa.get(col, ""))
            # Colorir linha em amarelo se for nova (adicionada hoje)
            cell.fill = PatternFill("solid", fgColor=COR_NOVA)
            if col == "link":
                cell.hyperlink = bolsa.get("link", "")
                cell.font = Font(color="0000EE", underline="single")

        ids_vistos.add(bid)
        adicionadas += 1

    # Após inserir, recolorir linhas pares normais (exceto as novas de hoje)
    hoje = str(date.today())
    for row in ws.iter_rows(min_row=2):
        data_cell = row[COLUNAS.index("data_coleta")]
        if data_cell.value != hoje:
            cor = COR_LINHA_PAR if row[0].row % 2 == 0 else "FFFFFF"
            for cell in row:
                if not isinstance(cell.fill, PatternFill) or cell.fill.fgColor.rgb == COR_NOVA:
                    cell.fill = PatternFill("solid", fgColor=cor)

    # Adiciona auto-filtro no cabeçalho
    ws.auto_filter.ref = ws.dimensions

    wb.save(ARQUIVO_EXCEL)
    log.info(f"Excel salvo: {ARQUIVO_EXCEL} ({adicionadas} novas bolsas)")
    return adicionadas


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("Início da coleta de bolsas")
    log.info("=" * 60)

    ids_vistos = carregar_ids_vistos()

    todas_bolsas = []

    coletores = [
        ("CNPq",        coletar_cnpq),
        ("CAPES",       coletar_capes),
        ("FAPESP",      coletar_fapesp),
        ("Fulbright",   coletar_fulbright),
        ("DAAD",        coletar_daad),
        ("RSS Genéricos", coletar_rss_genericos),
    ]

    for nome, func in coletores:
        try:
            log.info(f"Coletando: {nome}...")
            resultado = func()
            todas_bolsas.extend(resultado)
            time.sleep(2)  # pausa gentil entre fontes
        except Exception as e:
            log.error(f"Erro em {nome}: {e}")

    adicionadas = salvar_excel(todas_bolsas, ids_vistos)
    salvar_ids_vistos(ids_vistos)

    log.info(f"Coleta concluída. {adicionadas} novas bolsas adicionadas.")
    log.info(f"Total acumulado no banco: {len(ids_vistos)} bolsas únicas.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
